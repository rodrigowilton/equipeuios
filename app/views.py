import os
import subprocess
import logging
import tempfile
from django.shortcuts import render, redirect
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from .forms import MessageForm, AppBancoForm
from .models import AppBanco

# Configurar o logger
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)


def error_page(request):
	return render(request, 'error_page.html')


def send_message(request):
	if request.method == 'POST':
		form = MessageForm(request.POST, request.FILES)
		if form.is_valid():
			contacts = form.cleaned_data['contacts']
			message = form.cleaned_data['message']
			attachment = request.FILES.get('attachment')
			
			# Caminho absoluto para o script
			script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'send_whatsapp_message.py'))
			logger.debug(f"Script Path: {script_path}")
			
			# Verificar se o arquivo do script existe
			if not os.path.isfile(script_path):
				logger.error(f"Script file does not exist: {script_path}")
				return redirect('error_page')
			
			# Obter o caminho absoluto do executável Python
			python_executable = os.path.abspath(
				os.path.join(os.path.dirname(__file__), '..', 'venv', 'Scripts', 'python.exe'))
			logger.debug(f"Python Executable: {python_executable}")
			
			# Verificar se o Python executável existe
			if not os.path.isfile(python_executable):
				logger.error(f"Python executable does not exist: {python_executable}")
				return redirect('error_page')
			
			# Processar cada contato selecionado
			for contact in contacts:
				phone_number = contact.celular
				attachment_path = None
				
				if attachment:
					# Criar um arquivo temporário somente para leitura
					temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(attachment.name)[1])
					try:
						for chunk in attachment.chunks():
							temp_file.write(chunk)
					finally:
						temp_file.close()
					attachment_path = temp_file.name
					logger.debug(f"Attachment Path: {attachment_path}")
				
				# Verificar se o arquivo temporário realmente existe
				if attachment_path and not os.path.isfile(attachment_path):
					logger.error(f"Temporary file does not exist: {attachment_path}")
					continue
				
				# Preparar o comando para executar o script
				command = [python_executable, script_path, phone_number, message]
				if attachment_path:
					command.append(attachment_path)
				
				logger.debug(f"Command: {command}")
				
				# Chamar o script
				try:
					result = subprocess.run(command, check=True, text=True, capture_output=True)
					logger.debug(f"Output: {result.stdout}")
					logger.debug(f"Error: {result.stderr}")
				except subprocess.CalledProcessError as e:
					logger.error(f"Error occurred: {e}")
				finally:
					# Remover o arquivo temporário após o uso
					if attachment_path and os.path.isfile(attachment_path):
						os.remove(attachment_path)
			
			return redirect('success_page')
	else:
		form = MessageForm()
		contacts = AppBanco.objects.filter(aceita_whatsapp=True)
	
	return render(request, 'send_message.html', {'form': form, 'contatos': contacts})


def success_page(request):
	return render(request, 'success_page.html')

def import_success(request):
	return render(request, 'import_success_page.html')



def index(request):
	return render(request, "index.html")


def create_appbancouios(request):
	if request.method == 'POST':
		form = AppBancoForm(request.POST)
		if form.is_valid():
			try:
				form.save()
				return redirect('index')
			except Exception as e:
				print(f"Erro ao salvar o formulário: {e}")
		else:
			print(f"Formulário inválido: {form.errors}")
	else:
		form = AppBancoForm()
	
	return render(request, 'create_appbanco.html', {'form': form})


def export_to_excel(request):
	dados = AppBanco.objects.all()
	context = {
		'dados': dados
	}
	return render(request, 'export_to_excel.html', context)


def import_to_excel(request):
	if request.method == 'POST':
		file = request.FILES['file']
		if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
			# Salvar o arquivo temporariamente
			file_path = os.path.join(tempfile.gettempdir(), file.name)
			with open(file_path, 'wb+') as temp_file:
				for chunk in file.chunks():
					temp_file.write(chunk)
					
			# limpa o banco
			AppBanco.objects.all().delete()
			
			# Ler o arquivo
			wb = load_workbook(file_path, data_only=True)
			ws = wb.active
			
			# Obter os dados da planilha
			rows = ws.iter_rows(min_row=2, values_only=True)  # Supondo que a primeira linha são os cabeçalhos
			
			# Inserir os dados no banco de dados
			for row in rows:
				AppBanco.objects.create(
					nome=row[0],
					celular=row[1],
					cep=row[2],
					endereco=row[3],
					complemento=row[4],
					bairro=row[5],
					localidade=row[6],
					uf=row[7],
					numero=row[8],
					datacadastro=row[9],
					aceita_whatsapp=row[10]
				)
			
			# Deletar o arquivo temporário
			os.remove(file_path)
			
			return redirect('import_success_page')
		else:
			return HttpResponse('Formato de arquivo inválido.')
	
	return render(request, 'import_form.html')

def cliente_page(request):
    return render(request, 'cliente_page.html')
